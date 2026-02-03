import datetime
import io

import openpyxl
import pytest
from django.urls import reverse
from freezegun import freeze_time
from testutils.factories import (
    ContractFactory,
    ResourceFactory,
    SuperUserFactory,
    TaskFactory,
    TimeEntryFactory,
    WorkScheduleFactory,
)

from krm3.timesheet.report.payslip_report import report_timeentry_key_mapping
from tests.unit.web.test_views import _assert_homepage_content


@freeze_time('2025-08-22')
def test_report_view_current_month(client):
    SuperUserFactory(username='user00', password='pass123')
    resource = ResourceFactory()
    ContractFactory(resource=resource, work_schedule=WorkScheduleFactory())
    TimeEntryFactory(
        resource=resource,
        day_shift_hours=0,
        date=datetime.date.today(),
        holiday_hours=8,
    )
    client.login(username='user00', password='pass123')
    url = reverse('report')
    response = client.get(url)
    _assert_homepage_content(response)
    assert response.status_code == 200
    content = response.content.decode()
    assert f'{resource.last_name}</strong> {resource.first_name}' in content
    assert 'Report August 2025</h1>' in content


@freeze_time('2025-08-22')
@pytest.mark.parametrize(
    'month, expected_result',
    [
        pytest.param('202509', 'Report September 2025', id='next_month'),
        pytest.param('202507', 'Report July 2025', id='previous_month'),
    ],
)
def test_report_view_next_previous_month(client, month, expected_result):
    SuperUserFactory(username='user00', password='pass123')
    client.login(username='user00', password='pass123')
    response = client.get(reverse('report-month', args=[month]))
    _assert_homepage_content(response)
    assert response.status_code == 200
    assert expected_result in response.content.decode()


@pytest.mark.django_db
def test_report_creation(admin_client):  # noqa: PLR0915
    contract_1 = ContractFactory(work_schedule=WorkScheduleFactory())
    contract_2 = ContractFactory(work_schedule=WorkScheduleFactory())
    r1 = contract_1.resource
    r2 = contract_2.resource

    task_1 = TaskFactory(resource=r1)

    TimeEntryFactory(
        date=datetime.date(2025, 6, 5),
        day_shift_hours=8,
        task=task_1,
        resource=r1,
    )
    TimeEntryFactory(
        date=datetime.date(2025, 6, 6),
        day_shift_hours=6,
        task=task_1,
        resource=r1,
    )
    TimeEntryFactory(
        date=datetime.date(2025, 6, 9),
        day_shift_hours=0,
        sick_hours=8,
        protocol_number=12321,
        resource=r1,
    )
    TimeEntryFactory(
        date=datetime.date(2025, 6, 10),
        day_shift_hours=0,
        sick_hours=8,
        resource=r1,
    )
    TimeEntryFactory(
        date=datetime.date(2025, 6, 11),
        day_shift_hours=0,
        sick_hours=8,
        resource=r1,
    )

    TimeEntryFactory(
        date=datetime.date(2025, 6, 13),
        night_shift_hours=7,
        task=task_1,
        resource=r2,
    )

    date_arg = '202506'
    url = reverse('export_report', args=[date_arg])
    response = admin_client.get(url)
    assert response.status_code == 200
    assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    workbook = openpyxl.load_workbook(filename=io.BytesIO(response.content))
    r1_name = f'{r1.last_name.upper()} {r1.first_name}'
    r2_name = f'{r2.last_name.upper()} {r2.first_name}'
    assert len(workbook.sheetnames) == 1
    sheet_name = 'Resource report June 2025'
    assert sheet_name in workbook.sheetnames
    sheet = workbook[sheet_name]

    resource_rows = {}
    for row in range(1, sheet.max_row + 1):
        if cell_value := sheet[f'A{row}'].value:
            if r1_name in cell_value:
                resource_rows['r1'] = row
            elif r2_name in cell_value:
                resource_rows['r2'] = row

    assert 'r1' in resource_rows, f"Resource 1 '{r1_name}' not found in sheet"
    assert 'r2' in resource_rows, f"Resource 2 '{r2_name}' not found in sheet"

    r1_row = resource_rows['r1']
    r2_row = resource_rows['r2']

    assert sheet[f'A{r1_row}'].value.split(' - ')[-1] == r1_name
    assert sheet[f'A{r1_row + 1}'].value == 'Days 20'
    row_labels = [sheet[f'A{r1_row + 2 + i}'].value for i in range(9)]
    row_labels = [label for label in row_labels if label]
    assert all(
        value in report_timeentry_key_mapping.values()
        for value in [label for label in row_labels if label != 'Sick 12321']
    )

    first_resource_row = min(resource_rows.values())
    assert sheet[f'C{first_resource_row + 1}'].value == 'Sun\n1'
    assert sheet[f'AF{first_resource_row + 1}'].value == 'Mon\n30'
    assert sheet[f'AG{first_resource_row + 1}'].value is None

    regular_hours_data_row = r1_row + 2
    sick_hours_data_row = r1_row + 7
    nigh_shift_hours_data_row = r2_row + 3
    assert sheet[f'G{regular_hours_data_row}'].value == 8
    assert sheet[f'H{regular_hours_data_row}'].value == 6
    assert sheet[f'K{regular_hours_data_row}'].value is None
    assert sheet[f'B{regular_hours_data_row}'].value == 14
    assert sheet[f'L{sick_hours_data_row}'].value == 8
    assert sheet[f'K{sick_hours_data_row + 1}'].value == 8
    assert sheet[f'B{sick_hours_data_row}'].value == 16
    assert sheet[f'O{nigh_shift_hours_data_row}'].value == 7
