import datetime
import io
import json
import typing
from unittest.mock import patch
from constance.test import override_config

import openpyxl
import pytest
from bs4 import BeautifulSoup
from django.urls import reverse
from django.contrib.auth.models import Permission
from testutils.date_utils import _dt
from testutils.factories import (
    ProjectFactory,
    UserFactory,
    SuperUserFactory,
    ResourceFactory,
    TimeEntryFactory,
    SpecialLeaveReasonFactory,
    TaskFactory,
    ContractFactory,
)

from freezegun import freeze_time

from krm3.timesheet.report.payslip_report import report_timeentry_key_mapping
from tests._extras.testutils.factories import TimesheetSubmissionFactory

if typing.TYPE_CHECKING:
    from krm3.core.models import Resource


def _assert_homepage_content(response):
    assert response.status_code == 200
    content = response.content.decode()

    report_expected_url = reverse('report')
    task_report_expected_url = reverse('task_report')
    availability_report_expected_url = reverse('availability')
    releases_expected_url = reverse('releases')

    assert 'Report' in content
    assert 'Report by task' in content
    assert 'Availability report' in content
    assert f'href="{report_expected_url}"' in content
    assert f'href="{task_report_expected_url}"' in content
    assert f'href="{availability_report_expected_url}"' in content
    assert f'href="{releases_expected_url}"' in content


@pytest.mark.parametrize(
    'url', ('/be/', '/be/home/', '/be/availability/', '/be/releases/', '/be/report/', '/be/task_report/')
)
def test_resource_user_should_see_all_be_views(resource_client, url):
    response = resource_client.get(url)
    _assert_homepage_content(response)


@pytest.mark.parametrize(
    'url',
    [
        pytest.param('/be/report/', id='report'),
        pytest.param('/be/task_report/', id='task_report'),
    ],
)
def test_user_without_permission_should_only_see_its_reports(url, resource_client):
    another_user = UserFactory(username='user01', password='pass123')
    another_resource = ResourceFactory(user=another_user, profile=another_user.profile)
    response = resource_client.get(url)
    assert response.status_code == 200
    content = response.content.decode()
    resource_name = f'{resource_client._resource.last_name}</strong> {resource_client._resource.first_name}'
    another_user_name = f'{another_resource.last_name}</strong> {another_resource.first_name}'
    assert resource_name in content, f'{resource_name} not found in report'
    assert another_user_name not in content, f'{another_user_name} found in report'


@pytest.mark.parametrize('url', ('/be/report/', '/be/task_report/'))
def test_user_with_permissions_should_see_all_resources_reports(url, client):
    user = UserFactory(username='user00', password='pass123')
    another_user = UserFactory(username='user01', password='pass123')
    resource = ResourceFactory(user=user, profile=user.profile)
    another_resource = ResourceFactory(user=another_user, profile=another_user.profile)
    for perm in ['manage_any_timesheet', 'view_any_timesheet']:
        user.user_permissions.add(Permission.objects.get(codename=perm))
        client.login(username='user00', password='pass123')
        response = client.get(url)
        assert response.status_code == 200
        content = response.content.decode()
        resource_name = f'{resource.last_name}</strong> {resource.first_name}'
        another_user_name = f'{another_resource.last_name}</strong> {another_resource.first_name}'
        assert resource_name in content
        assert another_user_name in content


@pytest.mark.parametrize(
    'url', ('/be/home/', '/be/', '/be/availability/', '/be/report/', '/be/task_report/', '/be/releases/')
)
def test_not_authenticated_user_should_be_redirected_to_login_page(client, url):
    response = client.get(url)
    assert response.status_code == 302
    assert response.url == f'/admin/login/?next={url}'


@override_config(
    DEFAULT_RESOURCE_SCHEDULE=json.dumps({'mon': 8, 'tue': 8, 'wed': 8, 'thu': 8, 'fri': 8, 'sat': 8, 'sun': 8})
)
@freeze_time('2025-08-22')
def test_availability_view_current_month(client):
    SuperUserFactory(username='user00', password='pass123')
    resource = ResourceFactory()
    ContractFactory(resource=resource)
    TimeEntryFactory(
        resource=resource,
        day_shift_hours=0,
        date=datetime.date.today(),
        holiday_hours=8,
    )
    TimeEntryFactory(
        resource=resource,
        date=datetime.date.today() + datetime.timedelta(days=1),
        day_shift_hours=0,
        leave_hours=3,
        special_leave_hours=3,
        special_leave_reason=SpecialLeaveReasonFactory(),
    )
    TimeEntryFactory(
        resource=resource,
        date=datetime.date.today() + datetime.timedelta(days=2),
        day_shift_hours=0,
        rest_hours=4,
        leave_hours=4,
    )
    TimeEntryFactory(
        resource=resource,
        date=datetime.date.today() + datetime.timedelta(days=3),
        day_shift_hours=0,
        sick_hours=8,
    )
    TimeEntryFactory(
        resource=resource,
        date=datetime.date.today() + datetime.timedelta(days=4),
        day_shift_hours=0,
        leave_hours=3,
        special_leave_hours=3,
        rest_hours=2,
        special_leave_reason=SpecialLeaveReasonFactory(),
    )
    client.login(username='user00', password='pass123')
    url = reverse('availability')
    response = client.get(url)
    _assert_homepage_content(response)
    assert response.status_code == 200
    content = response.content.decode()
    assert f'{resource.first_name} {resource.last_name}' in content
    assert 'H' in content
    assert 'L 3.00, SL 3.00' in content
    assert 'L 4.00, R 4.00' in content
    assert 'S' in content
    assert 'L 3.00, SL 3.00, R 2.00' in content
    assert '<h1 class="title">Availability Report August 2025</h1>' in content


@freeze_time('2025-10-17')
def test_availability_view_filtered_by_project(client):
    SuperUserFactory(username='user00', password='pass123')
    project = ProjectFactory()
    resource = ResourceFactory()
    ContractFactory(resource=resource)
    TaskFactory(project=project, resource=resource)
    another_project = ProjectFactory()
    another_resource = ResourceFactory()
    ContractFactory(resource=another_resource)
    TaskFactory(project=another_project, resource=another_resource)
    TimeEntryFactory(
        resource=resource,
        day_shift_hours=0,
        date=_dt('20251017'),
        holiday_hours=8,
    )
    TimeEntryFactory(
        resource=another_resource,
        day_shift_hours=0,
        date=_dt('20251017'),
        leave_hours=3,
    )
    client.login(username='user00', password='pass123')
    url = reverse('availability')
    response = client.get(f'{url}?project={project.id}')
    _assert_homepage_content(response)
    assert response.status_code == 200
    content = response.content.decode()
    assert f'{resource.first_name} {resource.last_name}' in content
    assert f'{another_resource.first_name} {another_resource.last_name}' not in content
    assert 'H' in content
    assert 'L 3.00' not in content


@freeze_time('2025-08-22')
def test_availability_report_only_resources_with_contract(client):
    """Test that unprivileged user gets their own resource when project filter excludes them."""
    project = ProjectFactory()

    r_current: 'Resource' = ResourceFactory()
    ContractFactory(resource=r_current)
    r_past = ResourceFactory()
    ContractFactory(resource=r_past, period=(_dt('2020-01-01'), _dt('2021-01-01')))
    r_future = ResourceFactory()
    ContractFactory(resource=r_future, period=(_dt('2025-09-01'), _dt('2026-01-01')))

    TaskFactory(project=project, resource=r_current)
    TaskFactory(project=project, resource=r_past, end_date=_dt('20201231'))
    TaskFactory(project=project, resource=r_future)

    client.login(username=r_current.user.username, password=r_current.user._password)

    url = reverse('availability')
    response = client.get(f'{url}?project={project.id}')

    assert response.status_code == 200
    content = response.content.decode()

    assert f'{r_current.first_name} {r_current.last_name}' in content
    assert f'{r_past.first_name} {r_past.last_name}' not in content
    assert f'{r_future.first_name} {r_future.last_name}' not in content


@freeze_time('2025-08-22')
@pytest.mark.parametrize(
    'month, expected_result',
    [
        pytest.param('202509', 'Availability Report September 2025', id='next_month'),
        pytest.param('202507', 'Availability Report July 2025', id='previous_month'),
    ],
)
def test_availability_view_next_previous_month(client, month, expected_result):
    SuperUserFactory(username='user00', password='pass123')
    client.login(username='user00', password='pass123')
    response = client.get(reverse('availability-report-month', args=[month]))
    _assert_homepage_content(response)
    assert response.status_code == 200
    assert f'<h1 class="title">{expected_result}</h1>' in response.content.decode()


@freeze_time('2025-08-22')
def test_report_view_current_month(client):
    SuperUserFactory(username='user00', password='pass123')
    resource = ResourceFactory()
    TimeEntryFactory(
        resource=resource,
        day_shift_hours=0,
        date=datetime.date.today(),
        holiday_hours=8,
    )
    client.login(username='user00', password='pass123')
    url = reverse('report')
    response = client.get(url)
    _assert_homepage_content(response)
    assert response.status_code == 200
    content = response.content.decode()
    assert f'{resource.last_name}</strong> {resource.first_name}' in content
    assert 'Report August 2025</h1>' in content


@freeze_time('2025-08-22')
@pytest.mark.parametrize(
    'month, expected_result',
    [
        pytest.param('202509', 'Report September 2025', id='next_month'),
        pytest.param('202507', 'Report July 2025', id='previous_month'),
    ],
)
def test_report_view_next_previous_month(client, month, expected_result):
    SuperUserFactory(username='user00', password='pass123')
    client.login(username='user00', password='pass123')
    response = client.get(reverse('report-month', args=[month]))
    _assert_homepage_content(response)
    assert response.status_code == 200
    assert expected_result in response.content.decode()


@freeze_time('2025-08-22')
def test_task_report_view_current_month(client):
    SuperUserFactory(username='user00', password='pass123')
    task = TaskFactory()
    TimeEntryFactory(resource=task.resource, day_shift_hours=8, date=datetime.date.today(), task=task)
    client.login(username='user00', password='pass123')
    url = reverse('task_report')
    response = client.get(url)
    _assert_homepage_content(response)
    assert response.status_code == 200
    content = response.content.decode()
    assert '<h1 class="title">Task Report August 2025</h1>' in content


@freeze_time('2025-08-22')
@pytest.mark.parametrize(
    'month, expected_result',
    [
        pytest.param('202509', 'Task Report September 2025', id='next_month'),
        pytest.param('202507', 'Task Report July 2025', id='previous_month'),
    ],
)
def test_task_report_view_next_previous_month(client, month, expected_result):
    SuperUserFactory(username='user00', password='pass123')
    client.login(username='user00', password='pass123')
    response = client.get(reverse('task-report-month', args=[month]))
    _assert_homepage_content(response)
    assert response.status_code == 200
    assert expected_result in response.content.decode()


@pytest.mark.django_db
def test_report_creation(admin_client):  # noqa: PLR0915
    contract_1 = ContractFactory()
    contract_2 = ContractFactory()
    r1 = contract_1.resource
    r2 = contract_2.resource

    task_1 = TaskFactory(resource=r1)

    TimeEntryFactory(
        date=datetime.date(2025, 6, 5),
        day_shift_hours=8,
        task=task_1,
        resource=r1,
    )
    TimeEntryFactory(
        date=datetime.date(2025, 6, 6),
        day_shift_hours=6,
        task=task_1,
        resource=r1,
    )
    TimeEntryFactory(
        date=datetime.date(2025, 6, 9),
        day_shift_hours=0,
        sick_hours=8,
        protocol_number=12321,
        resource=r1,
    )
    TimeEntryFactory(
        date=datetime.date(2025, 6, 10),
        day_shift_hours=0,
        sick_hours=8,
        resource=r1,
    )
    TimeEntryFactory(
        date=datetime.date(2025, 6, 11),
        day_shift_hours=0,
        sick_hours=8,
        resource=r1,
    )

    TimeEntryFactory(
        date=datetime.date(2025, 6, 13),
        night_shift_hours=7,
        task=task_1,
        resource=r2,
    )

    date_arg = '202506'
    url = reverse('export_report', args=[date_arg])
    response = admin_client.get(url)
    assert response.status_code == 200
    assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    workbook = openpyxl.load_workbook(filename=io.BytesIO(response.content))
    r1_name = f'{r1.last_name.upper()} {r1.first_name}'
    r2_name = f'{r2.last_name.upper()} {r2.first_name}'
    assert len(workbook.sheetnames) == 1
    sheet_name = 'Resource report June 2025'
    assert sheet_name in workbook.sheetnames
    sheet = workbook[sheet_name]

    resource_rows = {}
    for row in range(1, sheet.max_row + 1):
        if cell_value := sheet[f'A{row}'].value:
            if r1_name in cell_value:
                resource_rows['r1'] = row
            elif r2_name in cell_value:
                resource_rows['r2'] = row

    assert 'r1' in resource_rows, f"Resource 1 '{r1_name}' not found in sheet"
    assert 'r2' in resource_rows, f"Resource 2 '{r2_name}' not found in sheet"

    r1_row = resource_rows['r1']
    r2_row = resource_rows['r2']

    assert sheet[f'A{r1_row}'].value.split(' - ')[-1] == r1_name
    assert sheet[f'A{r1_row + 1}'].value == 'Days 20'
    row_labels = [sheet[f'A{r1_row + 2 + i}'].value for i in range(9)]
    row_labels = [label for label in row_labels if label]
    assert all(
        value in report_timeentry_key_mapping.values()
        for value in [label for label in row_labels if label != 'Sick 12321']
    )

    first_resource_row = min(resource_rows.values())
    assert sheet[f'C{first_resource_row + 1}'].value == 'Sun\n1'
    assert sheet[f'AF{first_resource_row + 1}'].value == 'Mon\n30'
    assert sheet[f'AG{first_resource_row + 1}'].value is None

    regular_hours_data_row = r1_row + 2
    sick_hours_data_row = r1_row + 7
    nigh_shift_hours_data_row = r2_row + 3
    assert sheet[f'G{regular_hours_data_row}'].value == 8
    assert sheet[f'H{regular_hours_data_row}'].value == 6
    assert sheet[f'K{regular_hours_data_row}'].value is None
    assert sheet[f'B{regular_hours_data_row}'].value == 14
    assert sheet[f'L{sick_hours_data_row}'].value == 8
    assert sheet[f'K{sick_hours_data_row + 1}'].value == 8
    assert sheet[f'B{sick_hours_data_row}'].value == 16
    assert sheet[f'O{nigh_shift_hours_data_row}'].value == 7


@patch(
    'pathlib.Path.read_text',
    return_value="""## 1.5.33 (2025-09-10)
        ### Fix
        - update template

        ## 1.5.32 (2025-09-09)

        ### Feat
        - add commitizen setup

        ### Fix
        - update bump command with interactive mode""",
)
def test_releases_view_with_valid_markdown(mock_file, client):
    UserFactory(username='user00', password='pass123')
    client.login(username='user00', password='pass123')

    response = client.get(reverse('releases'))
    content = response.content.decode()

    assert response.status_code == 200
    assert '1.5.33' in content
    assert '1.5.32' in content
    assert 'update template' in content
    assert 'add commitizen setup' in content
    assert 'update bump command' in content
    assert 'Changelog' in content


@patch('pathlib.Path.read_text', side_effect=FileNotFoundError())
def test_releases_view_with_missing_file_should_show_error(mock_open_func, client):
    UserFactory(username='user00', password='pass123')
    client.login(username='user00', password='pass123')

    response = client.get(reverse('releases'))
    content = response.content.decode()

    assert response.status_code == 200
    assert 'text-gray-400' in content
    assert 'CHANGELOG.md file not found' in content


@patch('pathlib.Path.read_text', side_effect=PermissionError())
def test_releases_view_with_file_read_error(mock_open_func, client):
    UserFactory(username='user00', password='pass123')
    client.login(username='user00', password='pass123')

    response = client.get(reverse('releases'))
    content = response.content.decode()

    assert response.status_code == 200
    assert 'Changelog' in content
    assert 'Error parsing CHANGELOG.md' in content


def test_releases_view_uses_settings_changelog_path(client, settings, tmp_path):
    """Test that ReleasesView uses settings.CHANGELOG_PATH for locating the changelog file."""
    UserFactory(username='user00', password='pass123')
    client.login(username='user00', password='pass123')

    # Create a custom changelog file
    custom_changelog = tmp_path / 'custom_changelog.md'
    custom_changelog.write_text("""## 2.0.0 (2025-11-28)

### Feat
- Custom changelog content for testing

### Fix
- Verify settings.CHANGELOG_PATH is used
""")

    # Override the CHANGELOG_PATH setting
    settings.CHANGELOG_PATH = custom_changelog

    response = client.get(reverse('releases'))
    content = response.content.decode()

    assert response.status_code == 200
    assert '2.0.0' in content
    assert 'Custom changelog content for testing' in content
    assert 'Verify settings.CHANGELOG_PATH is used' in content


def test_payslip_report_with_timesheet_submissions(client):
    """Test payslip report includes timesheet submissions in coverage."""
    SuperUserFactory(username='user00', password='pass123')
    resource = ResourceFactory()

    TimesheetSubmissionFactory(resource=resource)

    client.login(username='user00', password='pass123')
    url = reverse('export_report', args=['202001'])
    response = client.get(url)

    assert response.status_code == 200


def test_user_profile_view_with_non_existent_user_id_returns_404(client):
    """Test that accessing a profile with a non-existent user ID returns 404."""
    UserFactory(username='user00', password='pass123')
    client.login(username='user00', password='pass123')

    url = reverse('user_resource', args=[9999])
    response = client.get(url)

    assert response.status_code == 404


def test_user_profile_view_user_without_resource_returns_404(client):
    """Test that accessing a profile for a user without an associated resource returns 404."""
    user = UserFactory(username='user00', password='pass123')
    client.login(username='user00', password='pass123')

    url = reverse('user_resource', args=[user.pk])
    response = client.get(url)

    assert response.status_code == 404


def test_user_profile_view_read_only_mode_for_other_user(client):
    """Test that viewing another user's profile shows read-only view without form or submit button."""
    # Create two users with resources
    viewer = UserFactory(username='viewer', password='pass123')
    ResourceFactory(user=viewer, first_name='ViewerFirst', last_name='ViewerLast')

    other_user = UserFactory(username='otheruser', password='pass456', email='other@example.com')
    other_resource = ResourceFactory(user=other_user, first_name='OtherFirst', last_name='OtherLast')

    # Login as viewer
    client.login(username='viewer', password='pass123')

    # Access other user's profile
    url = reverse('user_resource', args=[other_user.pk])
    response = client.get(url)

    assert response.status_code == 200

    # Parse HTML to verify read-only view
    soup = BeautifulSoup(response.content, 'html.parser')

    # Check that h1 contains the resource's full name
    h1 = soup.find('h1')
    assert h1 is not None
    assert f'{other_resource.first_name} {other_resource.last_name}' in h1.get_text()

    # Check that the paragraph after h1 contains the user's email
    h1_parent = h1.parent
    paragraphs = h1_parent.find_all('p')  # type: ignore
    assert len(paragraphs) > 0
    assert other_user.email in paragraphs[0].get_text()

    # Verify that the form does not exist
    form = soup.find('form', {'class': 'profile-form'})
    assert form is None

    # Verify that first_name and last_name input fields don't exist
    first_name_input = soup.find('input', {'name': 'first_name'})
    last_name_input = soup.find('input', {'name': 'last_name'})
    assert first_name_input is None
    assert last_name_input is None

    # Verify that submit button doesn't exist
    submit_button = soup.find('button', {'type': 'submit'})
    assert submit_button is None

    # Verify that Back button exists
    back_button = soup.find('a', string=lambda text: text and 'Back' in text)  # type: ignore
    assert back_button is not None


def test_user_profile_view_post_to_other_user_resource_returns_403(client):
    """Test that attempting to POST to update another user's resource returns 403 Forbidden."""
    # Create two users with resources
    attacker = UserFactory(username='attacker', password='pass123')
    ResourceFactory(user=attacker, first_name='AttackerFirst', last_name='AttackerLast')

    victim = UserFactory(username='victim', password='pass456')
    victim_resource = ResourceFactory(user=victim, first_name='VictimFirst', last_name='VictimLast')

    # Login as attacker
    client.login(username='attacker', password='pass123')

    # Attempt to POST to victim's profile
    url = reverse('user_resource', args=[victim.pk])
    response = client.post(
        url,
        {
            'first_name': 'HackedFirst',
            'last_name': 'HackedLast',
        },
    )

    # Assert that 403 Forbidden is returned
    assert response.status_code == 403

    # Refresh from database and verify that victim's resource was NOT changed
    victim_resource.refresh_from_db()
    assert victim_resource.first_name == 'VictimFirst'
    assert victim_resource.last_name == 'VictimLast'


def test_user_profile_view_post_with_empty_fields_shows_validation_errors(client):
    """Test that POSTing with empty first_name and last_name shows validation errors."""
    user = UserFactory(username='testuser', password='pass123')
    resource = ResourceFactory(user=user, first_name='OriginalFirst', last_name='OriginalLast')

    client.login(username='testuser', password='pass123')

    url = reverse('user_resource', args=[user.pk])
    response = client.post(
        url,
        {
            'first_name': '',
            'last_name': '',
        },
    )

    assert response.status_code == 200

    # Parse HTML to verify validation errors are displayed
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find error messages for first_name field
    first_name_errors = soup.find_all('span', {'class': 'profile-field-error'})

    # Convert all error text to a single string for easier assertion
    error_texts = [error.get_text() for error in first_name_errors]

    # Assert that validation errors are present
    assert len(error_texts) >= 2, 'Expected at least 2 validation errors (one for each field)'
    assert any('required' in error.lower() for error in error_texts), "Expected 'required' error message"

    # Verify that the resource was NOT updated
    resource.refresh_from_db()
    assert resource.first_name == 'OriginalFirst'
    assert resource.last_name == 'OriginalLast'


def test_user_profile_view_post_with_empty_body_shows_validation_errors(client):
    """Test that POSTing with an empty body (no data) shows validation errors."""
    user = UserFactory(username='testuser', password='pass123')
    resource = ResourceFactory(user=user, first_name='OriginalFirst', last_name='OriginalLast')

    client.login(username='testuser', password='pass123')

    url = reverse('user_resource', args=[user.pk])
    response = client.post(url, {})

    assert response.status_code == 200

    # Parse HTML to verify validation errors are displayed
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find all error messages
    error_spans = soup.find_all('span', {'class': 'profile-field-error'})

    # Convert all error text to a list
    error_texts = [error.get_text() for error in error_spans]

    # Assert that validation errors are present for both required fields
    assert len(error_texts) >= 2, 'Expected at least 2 validation errors (one for each required field)'
    assert any('required' in error.lower() for error in error_texts), "Expected 'required' error message"

    # Verify that the resource was NOT updated
    resource.refresh_from_db()
    assert resource.first_name == 'OriginalFirst'
    assert resource.last_name == 'OriginalLast'


def test_user_profile_view_get_shows_prepopulated_fields(client):
    """Test that GET request shows form fields prepopulated with current user/resource data."""
    user = UserFactory(username='testuser', password='pass123', email='test@example.com')
    ResourceFactory(user=user, first_name='TestFirst', last_name='TestLast')

    client.login(username='testuser', password='pass123')

    url = reverse('user_resource', args=[user.pk])
    response = client.get(url)

    assert response.status_code == 200

    # Parse HTML to check prepopulated field values
    soup = BeautifulSoup(response.content, 'html.parser')

    first_name_field = soup.find('input', {'name': 'first_name'})
    last_name_field = soup.find('input', {'name': 'last_name'})

    # Assert that fields exists
    assert first_name_field is not None
    assert last_name_field is not None

    # Assert that fields are prepopulated with correct values
    assert first_name_field['value'] == 'TestFirst'
    assert last_name_field['value'] == 'TestLast'


def test_user_profile_view_post_updates_all_fields(client):
    """Test that POSTing to the profile view updates all four fields."""
    user = UserFactory(username='oldusername', password='pass123', email='old@example.com')
    resource = ResourceFactory(user=user, first_name='OldFirstName', last_name='OldLastName')

    client.login(username='oldusername', password='pass123')

    url = reverse('user_resource', args=[user.pk])
    response = client.post(
        url,
        {
            'first_name': 'NewFirstName',
            'last_name': 'NewLastName',
        },
    )

    assert response.status_code == 200
    content = response.content.decode()

    # Assert success message is in the response
    assert 'Profile updated successfully' in content

    # Refresh from database to get updated values
    resource.refresh_from_db()

    # Assert all fields were updated
    assert resource.first_name == 'NewFirstName'
    assert resource.last_name == 'NewLastName'


def test_user_profile_view_displays_profile_picture_and_qr_code(client):
    """Test that profile picture and QR code are displayed when user has UserProfile with picture and Resource with
    vcard."""
    user = UserFactory(username='userwitpic', password='pass123', email='withpic@example.com')
    picture_url = 'http://www.example.com/picture.jpg'

    # Get the automatically created profile and update it
    user_profile = user.profile
    user_profile.picture = picture_url
    user_profile.save()

    # VCARD V3 example
    vcard_text = """
        BEGIN:VCARD
        VERSION:3.0
        FN:John Doe
        N:Doe;John;;;
        EMAIL:john.doe@example.com
        TEL:+1234567890
        END:VCARD
    """

    ResourceFactory(user=user, profile=user_profile, first_name='John', last_name='Doe', vcard_text=vcard_text)

    client.login(username='userwitpic', password='pass123')

    url = reverse('user_resource', args=[user.pk])
    response = client.get(url)

    assert response.status_code == 200

    # Parse HTML to check for profile picture and QR code
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find the img tag with class 'profile-picture'
    profile_picture_img = soup.find('img', {'class': 'profile-picture'})

    # Assert the img tag exists
    assert profile_picture_img is not None

    # Assert the src attribute contains the picture URL
    assert profile_picture_img['src'] == picture_url

    # Find the QR code wrapper div
    qr_wrapper = soup.find('div', {'class': 'profile-qr-wrapper'})

    # Assert the QR code wrapper exists
    assert qr_wrapper is not None

    # Find the QR code image container
    qr_image_div = soup.find('div', {'class': 'profile-qr-image'})

    # Assert the QR code image container exists
    assert qr_image_div is not None
