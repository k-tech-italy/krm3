from decimal import Decimal
from typing import override, Protocol

import openpyxl

from django.utils.translation import gettext as _
from krm3.core.models import Resource, User
from krm3.timesheet.report.base import TimesheetReport
from krm3.utils.numbers import safe_dec
from krm3.web.report_styles import (
    centered,
    header_alignment,
    header_fill,
    header_font,
    light_grey_fill,
    nwd_fill,
    thin_border,
)


def get_report_timeentry_key_mapping() -> dict[str, str]:
    return {
        'regular_hours': _('Regular hours'),
        'night_shift': _('Night shift'),
        'on_call': _('On call'),
        'holiday': _('Holiday'),
        'leave': _('Leave'),
        'sick': _('Sick'),
        'rest': _('Rest'),
        'overtime': _('Overtime'),
        'meal_voucher': _('Meal voucher'),
    }


report_timeentry_key_mapping = get_report_timeentry_key_mapping()


class StreamWriter(Protocol):
    def write(self, data) -> None: ...  # noqa: ANN001


class TimesheetReportExport(TimesheetReport):
    need = {'extra_holidays'}

    def write_excel(self, stream: StreamWriter, title: str) -> None:  # noqa: C901,PLR0912,PLR0915
        mapping = get_report_timeentry_key_mapping()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title

        ws.column_dimensions['A'].width = 30

        current_row = 1

        for idx, resource in enumerate(self.resources, 1):
            # spacing between employees
            if idx > 1:
                current_row += 2

            resources_report_days = self.calendars[resource.id]

            # Header section
            headers = [
                f'{idx} - {resource.last_name.upper()} {resource.first_name}',
                '',
                *['X' if kd.holiday else '' for kd in resources_report_days],
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = header_alignment

            current_row += 1

            # Days row
            working_days = sum([0 if kd.nwd else 1 for kd in resources_report_days])
            giorni = [
                _('Days {working_days}').format(working_days=working_days),
                _('Total HH'),
                *[f'{day.day_of_week_short_i18n}\n{day.day}' for day in resources_report_days],
            ]
            for col, giorno in enumerate(giorni):
                cell = ws.cell(row=current_row, column=col + 1, value=giorno)
                cell.alignment = header_alignment
                cell.fill = header_fill
                if col > 1 and resources_report_days[col - 2].nwd:
                    cell.fill = nwd_fill

            current_row += 1
            sick_days_with_protocol = {}

            if any(rkd.has_data for rkd in resources_report_days):
                dynamic_mapping = {}
                special_leave_days = {}
                for rkd in resources_report_days:
                    if rkd.data_special_leave_reason:
                        special_leave_days.setdefault(rkd.data_special_leave_reason.title, []).append(rkd)
                    if rkd.data_sick and rkd.data_protocol_number:
                        sick_days_with_protocol.setdefault(rkd.data_protocol_number, []).append(rkd)

                for key, label in mapping.items():
                    dynamic_mapping[key] = get_report_timeentry_key_mapping().get(label, label)
                    if key == 'leave' and special_leave_days:
                        for sl_title in special_leave_days:
                            dynamic_mapping[f'special_leave_{sl_title}'] = _('Special leave ({sl_title})').format(
                                sl_title=sl_title
                            )
                    if key == 'sick':
                        for protocol in sick_days_with_protocol:
                            dynamic_mapping[f'sick_days_{protocol}'] = _('Sick {protocol}').format(protocol=protocol)

                plain_sick_row: int = 0
                plain_sick_dec: float = Decimal(0)

                for lnum, (key, label) in enumerate(dynamic_mapping.items()):
                    rownum = current_row + lnum
                    if key == 'sick':
                        plain_sick_row = rownum

                    cell = ws.cell(row=rownum, column=1, value=get_report_timeentry_key_mapping().get(label, label))
                    if lnum % 2:
                        cell.fill = light_grey_fill

                    tot = None

                    for dd_num, rkd in enumerate(resources_report_days, 3):
                        if key == 'sick':
                            value = rkd.data_sick if (rkd.data_sick and not rkd.data_protocol_number) else None
                            plain_sick_dec += safe_dec(value)
                        elif key.startswith('sick_days_'):
                            protocol = key.replace('sick_days_', '')
                            value = rkd.data_sick if (rkd.data_sick and rkd.data_protocol_number == protocol) else None
                        elif key.startswith('special_leave_') and key != 'special_leave_hours':
                            reason_title = key.replace('special_leave_', '')
                            value = (
                                rkd.data_special_leave_hours
                                if rkd in special_leave_days.get(reason_title, [])
                                else None
                            )
                        else:
                            value = getattr(rkd, f'data_{key}')
                        cell = ws.cell(row=rownum, column=dd_num, value=value if value != 0 else None)
                        cell.alignment = centered
                        if rkd.nwd:
                            cell.fill = nwd_fill
                        elif lnum % 2:
                            cell.fill = light_grey_fill
                        if value is not None:
                            tot = safe_dec(tot) + safe_dec(value)
                    # TOT cell
                    cell = ws.cell(row=rownum, column=2, value='' if tot is None else tot)
                    cell.alignment = centered
                    if lnum % 2:
                        cell.fill = light_grey_fill
                current_row = rownum

            else:
                ws.cell(row=current_row, column=1, value=_('No data available'))
                current_row += 1

            # Deleting empty sick row if all sick days have protocol
            if sick_days_with_protocol and not plain_sick_dec:
                ws.delete_rows(plain_sick_row)
                current_row -= 1

        wb.save(stream)

    @override
    def _get_resources(self, user: User) -> list[Resource]:
        if user.has_any_perm('core.manage_any_timesheet', 'core.view_any_timesheet'):
            return [*Resource.objects.filter(preferred_in_report=True)]
        return [user.get_resource()]
