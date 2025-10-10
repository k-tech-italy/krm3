import datetime
import decimal
import json
import typing
from decimal import Decimal as D  # noqa: N817

import openpyxl
from constance import config
from dateutil.relativedelta import relativedelta
from django.contrib.auth import get_user_model

from krm3.config import settings
from krm3.core.models import Contract, ExtraHoliday, Resource, TimeEntry
from krm3.core.models.timesheets import TimesheetSubmission
from krm3.timesheet.rules import Krm3Day
from krm3.utils.dates import KrmCalendar, KrmDay, get_country_holidays
from krm3.utils.tools import format_data
from krm3.utils.numbers import normal, safe_dec
from krm3.web.report_styles import centered, header_alignment, header_fill, header_font, nwd_fill, thin_border, \
    light_grey_fill


class StreamWriter(typing.Protocol):
    def write(self, data) -> None: ...  # noqa: ANN001


User = get_user_model()

_fields = ['day_shift', 'night_shift', 'sick', 'holiday', 'leave', 'on_call', 'travel', 'rest', 'bank_from']


timeentry_key_mapping = {
    'bank': 'Banca ore',
    'day_shift': 'Ore ordinarie',
    'night_shift': 'Ore notturne',
    'on_call': 'Reperibilità',
    'travel': 'Viaggio',
    'holiday': 'Ferie',
    'leave': 'Permessi',
    'rest': 'Riposo',
    'sick': 'Malattia',
    'overtime': 'Ore straordinarie',
    'meal_voucher': 'Buoni pasto',
}


def enrich_with_resource_calendar(
    results: dict[Resource, dict], from_date: datetime.date, to_date: datetime.date
) -> None:
    for resource, stats in results.items():
        submitted_days = get_submitted_dates(from_date, to_date, resource)
        stats['days'] = [
            KrmDay(d, submitted=d.date in submitted_days) for d in KrmCalendar().iter_dates(from_date, to_date)
        ]


def enrich_with_meal_voucher(
    results: dict[Resource, dict], from_date: datetime.date, to_date: datetime.date
) -> dict[str, str]:
    """
    Add meal vaucher calculation to the results.

    Value is 1 if resource worked > min_meal_voucher as per Contract.meal_voucher schedule.

    Returns:
        Dictionary with meal voucher mapping for the keymap

    """
    calendar = KrmCalendar()
    meal_voucher_mapping = {}

    for resource, stats in results.items():
        days_interval = (to_date - from_date).days + 1
        stats['meal_voucher'] = [None] * days_interval

        contracts = resource.get_contracts(from_date, to_date)

        for day_index, krm_day in enumerate(calendar.iter_dates(from_date, to_date)):
            date = krm_day.date

            contract = resource.contract_for_date(contracts, date)
            if not contract:
                continue

            min_threshold = (
                krm_day.is_holiday(contract.country_calendar_code) and contract.meal_voucher.get('sun')
            ) or contract.meal_voucher.get(krm_day.day_of_week_short.lower())
            if min_threshold is None:
                continue

            total_worked_hours = (
                stats['day_shift'][day_index]
                + stats['night_shift'][day_index]
                + stats['travel'][day_index]
                + stats['bank_from'][day_index]
            )
            if total_worked_hours >= min_threshold:
                stats['meal_voucher'][day_index] = 1

    meal_voucher_mapping['meal_voucher'] = 'Buoni pasto'
    return meal_voucher_mapping


class UiElement:
    def __init__(self, **kwargs: dict) -> None:
        self.klass = None
        self.styles = None
        for k, v in kwargs.items():
            setattr(self, k, v)

    def render(self) -> str:
        if self.value:
            return str(self.value)
        return ''


class ReportCell(UiElement):
    def __init__(self, value: typing.Any, **kwargs: dict) -> None:
        super().__init__(**kwargs)
        self.value = value

    @property
    def negative(self) -> bool:
        """Return True if the cell value is negative."""
        return normal(self.value).startswith('-')

class ResourceCell(ReportCell):
    def render(self) -> str:
        return (
            f'{self.value["index"]} -'
            f' <strong>{self.value["resource"].last_name}</strong> {self.value["resource"].first_name}'
        )


class ReportRow(UiElement):
    def __init__(self, **kwargs: dict) -> None:
        super().__init__(**kwargs)
        self.cells: list[ReportCell] = []

    def add_cell(self, cell: ReportCell | typing.Any, **kwargs: dict) -> ReportCell:
        if not isinstance(cell, ReportCell):
            cell = ReportCell(cell, **kwargs)
        self.cells.append(cell)
        return cell


class ReportBlock(UiElement):
    def __init__(self, resource: Resource) -> None:
        super().__init__()
        self.rows: list[ReportRow] = []
        self.resource = resource

    def has_data(self) -> bool:
        return bool(len(self.rows) > 1)

    def add_row(self, row: ReportRow | None, **kwargs: dict) -> ReportRow:
        if row is None:
            row = ReportRow(**kwargs)
        self.rows.append(row)
        return row

    @property
    def width(self) -> int:
        return len(self.rows[0].cells) + 1


class TimesheetReport:
    def __init__(self, from_date: datetime.date, to_date: datetime.date, user: User) -> None:
        if user.has_any_perm('core.manage_any_timesheet', 'core.view_any_timesheet'):
            self.resources = Resource.objects.filter(preferred_in_report=True)
        else:
            self.resources = [user.get_resource()]
        resource_ids = {r.id for r in self.resources}

        self.from_date = from_date
        self.to_date = to_date

        self.default_schedule: dict[str, float] = json.loads(config.DEFAULT_RESOURCE_SCHEDULE)

        top_period = to_date + datetime.timedelta(days=1) if to_date else None

        self.time_entries: list[TimeEntry] = TimeEntry.objects.select_related('special_leave_reason').filter(
            date__gte=self.from_date, date__lte=self.to_date, resource_id__in=resource_ids
        )

        self.special_leave_reasons: dict[int, str] = {
            te.special_leave_reason_id: te.special_leave_reason.title
            for te in self.time_entries
            if te.special_leave_reason_id
        }

        self.submissions: dict[int, list[tuple[datetime.date, datetime.date]]] = {}
        for ts in TimesheetSubmission.objects.filter(
            resource_id__in=resource_ids, closed=True, period__overlap=(from_date, top_period)
        ):
            self.submissions.setdefault(ts.resource_id, []).append((ts.period.lower, ts.period.upper))

        self.country_codes: set[str] = {settings.HOLIDAYS_CALENDAR}
        self.resource_contracts: dict[int, list[Contract]] = {}

        for contract in list(
            Contract.objects.filter(period__overlap=(from_date, top_period), resource_id__in=resource_ids)
        ):
            self.resource_contracts.setdefault(contract.resource_id, []).append(contract)
            if contract.country_calendar_code and contract.country_calendar_code not in self.country_codes:
                self.country_codes.add(contract.country_calendar_code)

        self.extra_holidays = self._get_extra_holidays()

        self._holiday_cache = {}

        self.calendars: dict[int, list[Krm3Day]] = self._get_calendars()

    def _get_holiday(self, kd: 'KrmDay', country_calendar_code: str) -> bool:
        """Return True if the day is holiday."""
        if res := self._holiday_cache.get((kd.date, country_calendar_code)):
            return res
        if (eh := self.extra_holidays.get(kd)) and (
            country_calendar_code in eh or country_calendar_code.split('-')[0] in eh
        ):
            hol = True
        else:
            cal = get_country_holidays(country_calendar_code=country_calendar_code)
            hol = not cal.is_working_day(kd.date)
        return self._holiday_cache.setdefault((kd.date, country_calendar_code), hol)

    def _get_calendars(self) -> dict[int, list[Krm3Day]]:
        """Return the dict of KrmDay in the interval for the resource id.

        The KrmDay is enriched with:
        - min_working_hours: the float min number of working hours expected by the resource in the day
        - is_holiday: is overridden with a bool
        """
        ret: dict[int, list[Krm3Day]] = {}
        for resource in self.resources:
            res_id = resource.id
            contracts: list[Contract] = self.resource_contracts.get(res_id) or []

            ret[res_id] = list(Krm3Day(self.from_date, resource=resource).range_to(self.to_date))

            for kd in ret[res_id]:
                kd.resource = resource
                for c in contracts:
                    if c.falls_in(kd):
                        kd.contract = c
                        break

                country_calendar_code = (
                    kd.contract.country_calendar_code
                    if kd.contract and kd.contract.country_calendar_code
                    else settings.HOLIDAYS_CALENDAR
                )
                kd.holiday = self._get_holiday(kd, country_calendar_code)
                min_working_hours = self._get_min_working_hours(kd)
                kd.nwd = kd.contract is None or kd.holiday or min_working_hours == 0
                if not kd.nwd:
                    kd.min_working_hours = min_working_hours
                for p_lower, p_upper in self.submissions.get(res_id, []):
                    if p_lower <= kd.date < p_upper:
                        kd.submitted = True
                kd.apply([te for te in self.time_entries if te.resource_id == res_id and te.date == kd.date])

        return ret

    def _get_min_working_hours(self, kd: Krm3Day) -> float:
        """Return the minimum working hours for a given KrmDay enriched with the eventual contract.

        NB: the function will not consider if it is holiday. The check must be performed by the caller.
        """
        if kd.contract and kd.contract.working_schedule:
            schedule = kd.contract.working_schedule
        else:
            schedule = self.default_schedule
        return schedule[kd.day_of_week_short.lower()]

    def _get_extra_holidays(self) -> dict[KrmDay, list[str]]:
        """Retrieve the extra holidays for the given country codes."""
        short_codes = {x.split('-')[0] for x in self.country_codes}
        result = {}
        extra_holidays = list(
            ExtraHoliday.objects.filter(
                country_codes__overlap=list(self.country_codes.union(short_codes)),
                period__overlap=(self.from_date, self.to_date),
            )
        )

        for eh in extra_holidays:
            for kd in KrmDay(eh.period.lower).range_to(eh.period.upper - datetime.timedelta(days=1)):
                result.setdefault(kd, []).extend(eh.country_codes)
        return result

    def write_excel(self, stream: StreamWriter, title: str) -> None:  # noqa: C901,PLR0912
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title

        ws.column_dimensions['A'].width = 30

        current_row = 1

        for idx, resource in enumerate(self.resources, 1):
            # spacing between employees
            if idx > 1:
                current_row += 2

            resources_report_days = self.calendars[resource.id]

            # Header section
            headers = [
                f'{idx} - {resource.last_name.upper()} {resource.first_name}',
                '',
                *['X' if kd.holiday else '' for kd in resources_report_days],
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = header_alignment

            current_row += 1

            # Days row
            working_days = sum([0 if kd.nwd else 1 for kd in resources_report_days])
            giorni = [
                f'Giorni {working_days}',
                'Tot HH',
                *[
                    f'{day.day_of_week_short}\n{day.day}'
                    for day in resources_report_days
                ],
            ]
            for col, giorno in enumerate(giorni):
                cell = ws.cell(row=current_row, column=col + 1, value=giorno)
                cell.alignment = header_alignment
                cell.fill = header_fill
                if col > 2 and resources_report_days[col-2].nwd:
                    cell.fill = nwd_fill

            current_row += 1

            if any(rkd.has_data for rkd in resources_report_days):
                for lnum, (key, label) in enumerate(timeentry_key_mapping.items()):
                    if key != 'bank':
                        rownum = current_row + lnum

                        cell = ws.cell(row=rownum, column=1, value=label)
                        if lnum % 2:
                            cell.fill = light_grey_fill
                        tot = None

                        for dd_num, rkd in enumerate(resources_report_days, 3):
                            value = getattr(rkd, f'data_{key}')
                            cell = ws.cell(row=rownum, column=dd_num, value=value)
                            cell.alignment = centered
                            if lnum % 2:
                                cell.fill = light_grey_fill
                            if rkd.nwd:
                                cell.fill = nwd_fill
                            if value is not None:
                                tot = safe_dec(tot) + safe_dec(value)
                        cell = ws.cell(row=rownum, column=2, value='' if tot is None else tot)
                        cell.alignment = centered
                        if lnum % 2:
                            cell.fill = light_grey_fill
                    else:
                        current_row -= 1
                current_row = rownum

            else:
                ws.cell(row=current_row, column=1, value='No data available')
                current_row += 1

        wb.save(stream)

    def report_html(self) -> list[ReportBlock]:
        blocks = []
        for resource in self.resources:
            blocks.append(block := ReportBlock(resource))
            row = block.add_row(ReportRow())
            resources_report_days = self.calendars[resource.id]
            row.add_cell(sum([0 if kd.nwd else 1 for kd in resources_report_days]))
            for kd in resources_report_days:
                row.add_cell(kd)

            if any(rkd.has_data for rkd in resources_report_days):
                for key, label in timeentry_key_mapping.items():
                    row = block.add_row(ReportRow())
                    row.add_cell(label)
                    row.add_cell(cell_tot_hh := ReportCell(decimal.Decimal(0)))
                    for rkd in resources_report_days:
                        value = getattr(rkd, f'data_{key}')
                        row.add_cell(normal(value)).nwd = rkd.nwd
                        cell_tot_hh.value += value or decimal.Decimal(0)
                    cell_tot_hh.value = normal(cell_tot_hh.value)

                # special leaves
                special_leave_days: dict[str, list[Krm3Day]] = {}
                for rkd in resources_report_days:
                    if rkd.data_special_leave_reason:
                        special_leave_days.setdefault(rkd.data_special_leave_reason, []).append(rkd)

                for sl_id, sl_days in special_leave_days.items():
                    row = ReportRow()
                    row.add_cell(f'Perm. speciale ({sl_id})')
                    row.add_cell(cell_tot_hh := ReportCell(decimal.Decimal(0)))
                    for rkd in resources_report_days:
                        value = rkd.data_special_leave_hours if rkd in sl_days else ''
                        row.add_cell(normal(value)).nwd = rkd.nwd
                        cell_tot_hh.value += value or decimal.Decimal(0)
                    cell_tot_hh.value = normal(cell_tot_hh.value)
                    block.rows.insert(len(block.rows) - 4, row)

        return blocks


def timesheet_report_raw_data(
    from_date: datetime.date, to_date: datetime.date, resource: Resource | None = None
) -> tuple[dict['Resource', dict[str, list[D]]], dict[str, str]]:
    """
    Prepare the data for the timesheet report.

    If the resource is not provided, the report will be for all resources.

    Returns:
        A dictionary mapping resource names to a dict of time_entry_types with a list of
        Decimals summing up hours for each day.

    """
    qs = TimeEntry.objects.filter(date__gte=from_date, date__lte=to_date, resource__active=True).order_by(
        'resource', 'date'
    )

    if resource:
        qs = qs.filter(resource=resource)

    start_date = KrmDay(from_date)
    days_interval = (to_date - from_date).days + 1
    results, special_leave_mapping_dict = {}, {}
    for entry in qs:
        date = KrmDay(entry.date)
        resource_stats = results.setdefault(entry.resource, {})

        index = date - start_date
        if entry.special_leave_reason:
            key = f'special_leave|{entry.special_leave_reason.title}'
            special_leave_mapping_dict[key] = f'Perm. speciale ({entry.special_leave_reason.title})'
            hours_types = resource_stats.setdefault(key, [D('0.00')] * days_interval)
            hours_types[index] += entry.special_leave_hours
        for field in _fields:
            hours_types = resource_stats.setdefault(field, [D('0.00')] * days_interval)
            if field == 'bank_from':
                hours_types[index] += getattr(entry, f'{field}')
            else:
                hours_types[index] += getattr(entry, f'{field}_hours')

    for stats in results.values():
        stats['overtime'] = [D('0.00')] * days_interval

    enrich_with_resource_calendar(results, from_date, to_date)
    meal_voucher_mapping = enrich_with_meal_voucher(results, from_date, to_date)
    additional_mapping = special_leave_mapping_dict | meal_voucher_mapping
    return results, additional_mapping


def add_report_summaries(results: dict) -> None:
    """Insert summary fields to the results at the beginning of each list."""
    for stats in results.values():
        for k, result_list in stats.items():
            if k != 'days':
                if k == 'meal_voucher':
                    voucher_count = sum(1 for val in result_list if val == 1)
                    result_list.insert(0, voucher_count)
                else:
                    result_list.insert(0, sum(result_list))
                for i in range(1, len(result_list)):
                    if result_list[i] == D('0.00'):
                        result_list[i] = None


def calculate_overtime(resource_stats: dict) -> None:
    """Calculate overtime for each day."""
    for stats in resource_stats.values():
        num_days = len(stats['day_shift'])

        day_keys = [
            x
            for x in stats
            if x
            not in [
                'day_shift',
                'night_shift',
                'on_call',
                'travel',
                'rest',
                'overtime',
                'days',
                'bank_from',
                'bank_to',
            ]
        ]

        for i in range(num_days):
            if sum([stats[x][i] or D(0) for x in day_keys]) == D(0.0):
                tot_hours = stats['day_shift'][i] + stats['night_shift'][i] + stats['travel'][i]
                day_shift = min(D('8.00'), stats['day_shift'][i] + stats['night_shift'][i])
                stats['day_shift'][i] = day_shift
                stats['overtime'][i] = max(tot_hours - D('8.00'), D('0.00'))


def get_submitted_dates(from_date: datetime.date, to_date: datetime.date, resource: 'Resource') -> set[datetime.date]:
    calendar = KrmCalendar()
    submissions = TimesheetSubmission.objects.get_closed_in_period(from_date, to_date, resource).values('period')
    submitted_dates = set()

    for submission in submissions:
        period_start = submission['period'].lower
        period_end = submission['period'].upper - datetime.timedelta(days=1)
        actual_start = max(period_start, from_date)
        actual_end = min(period_end, to_date)

        submitted_dates.update(krm_day.date for krm_day in calendar.iter_dates(actual_start, actual_end))

    return submitted_dates


def get_days_submission(
    from_date: datetime.date, to_date: datetime.date, resource: Resource
) -> dict[datetime.date, bool]:
    """Return dictionary of all days in a period with their submission status for a specific resource."""
    calendar = KrmCalendar()

    submitted_days = get_submitted_dates(from_date, to_date, resource)

    return {krm_day.date: krm_day.date in submitted_days for krm_day in calendar.iter_dates(from_date, to_date)}


def timesheet_report_data(current_month: str | None, user: User) -> dict[str, typing.Any]:
    """Prepare the data for the timesheet report."""
    if current_month is None:
        start_of_month = datetime.date.today().replace(day=1)
    else:
        start_of_month = datetime.datetime.strptime(current_month, '%Y%m').date()
    prev_month = start_of_month - relativedelta(months=1)
    next_month = start_of_month + relativedelta(months=1)

    end_of_month = start_of_month + relativedelta(months=1, days=-1)
    resource = None
    if not user.has_any_perm('core.manage_any_timesheet', 'core.view_any_timesheet'):
        resource = user.get_resource()
    data, additional_mapping = timesheet_report_raw_data(start_of_month, end_of_month, resource=resource)
    calculate_overtime(data)
    add_report_summaries(data)

    for shifts in data.values():
        for key, values in shifts.items():
            if key != 'days':
                shifts[key] = [format_data(v) for v in values]
    resources = Resource.objects.filter(active=True)
    if resource:
        resources = resources.filter(pk=resource.pk)
    data = dict.fromkeys(resources.order_by('last_name', 'first_name'), None) | data
    days = list(Krm3Day(start_of_month.strftime('%Y-%m-%d')).range_to(end_of_month))

    return {
        'prev_month': prev_month.strftime('%Y%m'),
        'current_month': start_of_month.strftime('%Y%m'),
        'next_month': next_month.strftime('%Y%m'),
        'title': start_of_month.strftime('%B %Y'),
        'days': days,
        'data': data,
        'keymap': timeentry_key_mapping | additional_mapping,
    }
