from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

header_font = Font(name='Calibri', bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
centered = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
)
nwd_fill = PatternFill(start_color='979797', end_color='979797', fill_type='solid')
light_grey_fill = PatternFill(start_color='dddddd', end_color='dddddd', fill_type='solid')
