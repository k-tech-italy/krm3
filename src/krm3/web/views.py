import base64
import datetime
import json
import logging
import typing
from pathlib import Path
from typing import Any, cast, override

import binascii
import openpyxl
import markdown
from bs4 import BeautifulSoup
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Min
from django.http import HttpRequest, HttpResponse, Http404, HttpResponseBase
from django.shortcuts import get_object_or_404
from django.urls import reverse
from django.utils.text import slugify
from django.views.generic import TemplateView
from django.utils.translation import gettext_lazy as _
from django_simple_dms.models import Document, DocumentTag

from krm3.core.forms import ResourceForm
from krm3.core.models.projects import Project
from krm3.timesheet.report.availability import AvailabilityReportOnline
from krm3.timesheet.report.payslip import TimesheetReportOnline
from krm3.timesheet.report.payslip_report import TimesheetReportExport
from krm3.timesheet.report.task import TimesheetTaskReportOnline
from krm3.web.document_filter import DocumentFilter
from krm3.web.report_styles import centered, header_alignment, header_fill, header_font, nwd_fill, thin_border

if typing.TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from krm3.core.models import Contract, User as UserType

logger = logging.getLogger(__name__)


User = get_user_model()


class ReportMixin:
    def get_context_data(self, **kwargs) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['nav_bar_items'] = {
            'Report': reverse('report'),
            'Report by task': reverse('task_report'),
            'Availability report': reverse('availability'),
            'Releases': reverse('releases'),
        }
        context['logout_url'] = reverse('logout')

        return context


class HomeView(LoginRequiredMixin, TemplateView):
    login_url = '/admin/login/'
    template_name = 'home.html'

    def get_context_data(self, **kwargs) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['nav_bar_items'] = {
            'Report': reverse('report'),
            'Report by task': reverse('task_report'),
            'Availability report': reverse('availability'),
            'Releases': reverse('releases'),
        }
        context['logout_url'] = reverse('logout')

        return context


class UserResourceView(LoginRequiredMixin, ReportMixin, TemplateView):
    login_url = '/admin/login/'
    template_name = 'user_resource.html'

    @override
    def dispatch(self, request: HttpRequest, *args, **kwargs) -> HttpResponseBase:
        """Check permissions before processing the request."""
        user_id = kwargs.get('pk')
        user = cast('UserType', get_object_or_404(User, pk=user_id))

        # Check if user has an associated resource
        resource = user.get_resource()
        if not resource:
            raise Http404('No resource found for this user.')

        self.user = user
        self.resource = resource
        # Determine if the view should be read-only (viewing someone else's profile)
        self.read_only = request.user.pk != user_id
        return super().dispatch(request, *args, **kwargs)

    def post(self, request: HttpRequest, *args, **kwargs) -> HttpResponse:
        """Handle form submission."""
        # Prevent editing if viewing in read-only mode
        if self.read_only:
            raise PermissionDenied("You don't have permission to edit this profile.")

        form = ResourceForm(resource=self.resource, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _('Profile updated successfully.'))
            # Redirect to the same profile page after successful save
            return self.get(request, *args, **kwargs)

        # If form is invalid, re-render with errors
        context = self.get_context_data(**kwargs)
        context['form'] = form
        return self.render_to_response(context)

    def get_context_data(self, **kwargs) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)

        # Create form with existing data if not already in context (from POST)
        if 'form' not in context:
            context['form'] = ResourceForm(resource=self.resource)

        # Get profile picture URL if it exists
        profile_picture = None
        if hasattr(self.user, 'profile') and self.user.profile.picture:
            profile_picture = self.user.profile.picture

        # Get vcard_text if it exists and is not empty
        vcard_text = None
        if self.resource.vcard_text:
            vcard_text = self.resource.vcard_text

        context['resource'] = self.resource
        context['user'] = self.user
        context['profile_picture'] = profile_picture
        context['vcard_text'] = vcard_text
        context['read_only'] = self.read_only

        return context


class ScanQRView(TemplateView):
    template_name = 'scan_qr.html'


class AvailabilityReportView(LoginRequiredMixin, ReportMixin, TemplateView):
    login_url = '/admin/login/'
    template_name = 'availability_report.html'

    def get(self, request: HttpRequest, *args, month: str | None = None, **kwargs) -> HttpResponse:
        self.month = month
        return super().get(request, *args, **kwargs)

    def _get_base_context(self) -> dict:
        if self.month is None:
            start_of_month = datetime.date.today().replace(day=1)
        else:
            start_of_month = datetime.datetime.strptime(self.month, '%Y%m').date()

        prev_month = start_of_month - relativedelta(months=1)
        next_month = start_of_month + relativedelta(months=1)

        return {
            'start': start_of_month,
            'end': start_of_month + relativedelta(months=1, days=-1),
            'current_month': start_of_month.strftime('%Y%m'),
            'prev_month': prev_month.strftime('%Y%m'),
            'next_month': next_month.strftime('%Y%m'),
            'title': {
                'January': _('January'),
                'February': _('February'),
                'March': _('March'),
                'April': _('April'),
                'May': _('May'),
                'June': _('June'),
                'July': _('July'),
                'August': _('August'),
                'September': _('September'),
                'October': _('October'),
                'November': _('November'),
                'December': _('December'),
            }.get(start_of_month.strftime('%B'), '')
            + f'{start_of_month.strftime(" %Y")}',
        }

    def get_context_data(self, **kwargs) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        ctx = self._get_base_context()
        context.update(ctx)

        selected_project = self.request.GET.get('project', '')
        project_param = selected_project if selected_project else None

        projects = {'': _('All projects')} | dict(Project.objects.values_list('id', 'name'))
        context['projects'] = projects
        context['selected_project'] = selected_project
        report_blocks = AvailabilityReportOnline(
            ctx['start'], ctx['end'], cast('UserType', self.request.user), project_param
        )
        context['report_blocks'] = report_blocks.report_html()

        return context


def _write_resource_data(ws: 'Worksheet', data: dict, report_data: dict, current_row: int) -> int:
    """Write all data rows for a resource."""
    if not data:
        return current_row

    for key, value in data.items():
        if key in report_data['keymap']:
            safe_value = ['' if v is None else v for v in value]
            row_data = [report_data['keymap'][key], *safe_value]

            for col, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=cell_value)
                if col > 1:
                    cell.alignment = centered
            current_row += 1

    return current_row


def export_report(request: HttpRequest, report_data: dict, date: str) -> HttpResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Report {date[0:4]}-{date[4:6]}'

    ws.column_dimensions['A'].width = 30

    current_row = 1

    for resource_idx, (resource, data) in enumerate(report_data['data'].items(), start=1):
        if data is None:
            continue

        # spacing between employees
        if resource_idx > 1:
            current_row += 2

        holidays = []
        overlapping_contracts: 'list[Contract]' = resource.get_contracts(min(data['days']).date, max(data['days']).date)

        for day in data['days']:
            contract = resource.contract_for_date(overlapping_contracts, day)
            calendar_code = contract.country_calendar_code if contract else None
            holidays.append('X' if day.is_holiday(calendar_code) else '')

        # Header section
        headers = [
            f'{resource_idx} - {resource.last_name.upper()} {resource.first_name}',
            'Tot HH',
            *holidays,
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = header_alignment

        current_row += 1

        # Days row
        giorni = [
            'Giorni',
            '',
            *[
                f'{"**" if not day.submitted else ""}{day.day_of_week_short}\n{day.day}'
                f'{"**" if not day.submitted else ""}'
                for day in data['days']
            ],
        ]
        for col, giorno in enumerate(giorni):
            cell = ws.cell(row=current_row, column=col + 1, value=giorno)
            cell.alignment = header_alignment
            if col > 1 and (data['days'][col - 2].is_holiday() or data['days'][col - 2].min_working_hours == 0):
                cell.fill = nwd_fill

        current_row += 1

        current_row = _write_resource_data(ws, data, report_data, current_row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'report_{date[0:4]}-{date[4:6]}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


class ReportView(LoginRequiredMixin, ReportMixin, TemplateView):
    login_url = '/admin/login/'
    template_name = 'report.html'

    def get(
        self, request: HttpRequest, *args, month: str | None = None, export: bool = False, **kwargs
    ) -> HttpResponse:
        self.month = month
        if export:
            ctx = self._get_base_context()
            title = ctx['title']
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f'report_{slugify(title)}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            report = TimesheetReportExport(ctx['start'], ctx['end'], cast('UserType', self.request.user))

            # FIXME: `response` is incompatible with `payslip_report.StreamWriter`
            report.write_excel(response, _('Resource report {title}').format(title=title))
            return response
        return super().get(request, *args, **kwargs)

    def _get_base_context(self) -> dict:
        if self.month is None:
            start_of_month = datetime.date.today().replace(day=1)
        else:
            start_of_month = datetime.datetime.strptime(self.month, '%Y%m').date()

        prev_month = start_of_month - relativedelta(months=1)
        next_month = start_of_month + relativedelta(months=1)

        return {
            'start': start_of_month,
            'end': start_of_month + relativedelta(months=1, days=-1),
            'current_month': start_of_month.strftime('%Y%m'),
            'prev_month': prev_month.strftime('%Y%m'),
            'next_month': next_month.strftime('%Y%m'),
            'title': {
                'January': _('January'),
                'February': _('February'),
                'March': _('March'),
                'April': _('April'),
                'May': _('May'),
                'June': _('June'),
                'July': _('July'),
                'August': _('August'),
                'September': _('September'),
                'October': _('October'),
                'November': _('November'),
                'December': _('December'),
            }.get(start_of_month.strftime('%B'), '')
            + f'{start_of_month.strftime(" %Y")}',
        }

    def get_context_data(self, **kwargs) -> dict[str, Any]:
        ctx = super().get_context_data(**kwargs) | self._get_base_context()

        report_blocks = TimesheetReportOnline(ctx['start'], ctx['end'], cast('UserType', self.request.user))
        ctx['report_blocks'] = report_blocks.report_html()
        return ctx


class TaskReportView(LoginRequiredMixin, ReportMixin, TemplateView):
    login_url = '/admin/login/'
    template_name = 'task_report.html'

    def get(
        self, request: HttpRequest, *args, month: str | None = None, export: bool = False, **kwargs
    ) -> HttpResponse:
        self.month = month
        return super().get(request, *args, **kwargs)

    def _get_base_context(self) -> dict:
        if self.month is None:
            start_of_month = datetime.date.today().replace(day=1)
        else:
            start_of_month = datetime.datetime.strptime(self.month, '%Y%m').date()

        prev_month = start_of_month - relativedelta(months=1)
        next_month = start_of_month + relativedelta(months=1)

        return {
            'start': start_of_month,
            'end': start_of_month + relativedelta(months=1, days=-1),
            'current_month': start_of_month.strftime('%Y%m'),
            'prev_month': prev_month.strftime('%Y%m'),
            'next_month': next_month.strftime('%Y%m'),
            'title': {
                'January': _('January'),
                'February': _('February'),
                'March': _('March'),
                'April': _('April'),
                'May': _('May'),
                'June': _('June'),
                'July': _('July'),
                'August': _('August'),
                'September': _('September'),
                'October': _('October'),
                'November': _('November'),
                'December': _('December'),
            }.get(start_of_month.strftime('%B'), '')
            + f'{start_of_month.strftime(" %Y")}',
        }

    def get_context_data(self, **kwargs) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        ctx = self._get_base_context()
        context.update(ctx)

        report_blocks = TimesheetTaskReportOnline(ctx['start'], ctx['end'], cast('User', self.request.user))
        context['report_blocks'] = report_blocks.report_html()

        return context


class ReleasesView(HomeView):
    template_name = 'releases.html'

    def get_context_data(self, **kwargs) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        changelog_file_path = Path(settings.CHANGELOG_PATH)
        changelog_html = ''

        try:
            changelog_content = Path(changelog_file_path).read_text(encoding='utf-8')
            changelog_html = markdown.markdown(changelog_content)

            soup = BeautifulSoup(changelog_html, 'html.parser')
            for h2 in soup.find_all('h2'):
                # FIXME: `element.get()` does not necessarily return a list
                h2['class'] = h2.get('class', []) + [
                    'text-2xl',
                    'text-blue-300',
                    'border-b',
                    'border-white/20',
                    'pb-2',
                    'mb-4',
                    'mt-6',
                    'font-semibold',
                ]

            for h3 in soup.find_all('h3'):
                h3['class'] = h3.get('class', []) + ['text-xl', 'text-purple-300', 'mb-3', 'font-medium']

            for ul in soup.find_all('ul'):
                ul['class'] = ul.get('class', []) + ['space-y-2', 'my-4']

            for li in soup.find_all('li'):
                li['class'] = li.get('class', []) + ['marker:text-blue-400', 'marker:font-bold', 'ml-4']

            for p in soup.find_all('p'):
                p['class'] = p.get('class', []) + ['text-gray-200', 'leading-relaxed']

            for strong in soup.find_all('strong'):
                strong['class'] = strong.get('class', []) + ['text-white', 'font-semibold']

            changelog_html = str(soup)
        except FileNotFoundError:
            logger.warning(f'CHANGELOG.md file not found at {changelog_file_path}')
            changelog_html = "<p class='text-gray-400'>CHANGELOG.md file not found.</p>"
        except (OSError, UnicodeDecodeError) as e:
            logger.error(f'Error parsing CHANGELOG.md: {e}')
            changelog_html = f"<p class='text-red-400'>Error parsing CHANGELOG.md: {e}</p>"

        context['changelog_html'] = changelog_html
        return context


class DocumentListView(LoginRequiredMixin, ReportMixin, TemplateView):
    login_url = '/admin/login/'
    template_name = 'document_list.html'

    def get_template_names(self) -> list[str]:
        """Return partial template for HTMX requests."""
        if self.request.headers.get('HX-Request'):
            return ['partials/document_table.html']
        # self.template_name is Optional[str] on the base class; cast to str to satisfy the return type
        return [str(self.template_name)] if self.template_name else []

    def _get_base_queryset(self) -> Any:
        """Get base queryset of documents accessible by the current user."""
        try:
            return Document.objects.accessible_by(self.request.user).prefetch_related('tags')  # type: ignore
        except Exception as e:  # noqa: BLE001
            logger.error(f'Error getting accessible documents: {e}')
            messages.error(self.request, _('Error loading documents. Please try again.'))
            return Document.objects.none()

    def _parse_and_apply_filter(self, queryset: Any) -> tuple[Any, dict[str, Any] | None]:
        """Parse and apply filter from query string if present."""
        filter_b64 = self.request.GET.get('filter')
        if not filter_b64:
            return queryset, None

        current_filter = None
        try:
            filter_json = base64.b64decode(filter_b64).decode('utf-8')
            current_filter = json.loads(filter_json)
            logger.info(f'Parsed filter: {current_filter}')

            document_filter = DocumentFilter(queryset, current_filter)
            queryset = document_filter.apply()

            filter_errors = document_filter.get_errors()
            if filter_errors:
                for error in filter_errors:
                    messages.error(self.request, error)

        except json.JSONDecodeError as e:
            logger.warning(f'Invalid JSON in filter: {e}')
            messages.error(self.request, _('Invalid filter format. Please try again.'))
            current_filter = None
        except (UnicodeDecodeError, binascii.Error) as e:
            # Must come before ValueError since these are subclasses of ValueError
            logger.warning(f'Invalid base64 encoding in filter: {e}')
            messages.error(self.request, _('Invalid filter encoding. Please try again.'))
            current_filter = None
        except ValueError as e:
            # Catch ValueError from DocumentFilter (unsupported operators/fields)
            logger.warning(f'Invalid filter configuration: {e}')
            messages.error(self.request, str(e))
            current_filter = None
        except Exception as e:  # noqa: BLE001
            logger.error(f'Unexpected error applying filter: {e}')
            messages.error(self.request, _('Error applying filter. Please check your filter criteria.'))
            current_filter = None

        return queryset, current_filter

    def _apply_sorting(self, queryset: Any) -> tuple[Any, str]:
        """Apply sorting to queryset based on sort parameter."""
        sort_param = self.request.GET.get('sort', '-upload_date')
        valid_sort_fields = {
            'document': 'document',
            '-document': '-document',
            'upload_date': 'upload_date',
            '-upload_date': '-upload_date',
            'reference_period': 'reference_period',
            '-reference_period': '-reference_period',
            'tags': 'first_tag',
            '-tags': '-first_tag',
        }

        if sort_param in valid_sort_fields:
            order_field = valid_sort_fields[sort_param]
            if 'first_tag' in order_field:
                queryset = queryset.annotate(first_tag=Min('tags__title'))
            queryset = queryset.order_by(order_field)
        else:
            queryset = queryset.order_by('-upload_date')

        return queryset, sort_param

    def _paginate_queryset(self, queryset: Any) -> Any:
        """Paginate the queryset."""
        page_number = self.request.GET.get('page', 1)
        paginator = Paginator(queryset, 10)

        try:
            return paginator.page(page_number)
        except PageNotAnInteger:
            return paginator.page(1)
        except EmptyPage:
            return paginator.page(paginator.num_pages)

    def _get_available_tags(self) -> list[str]:
        """Get list of available tags for filter dropdown."""
        try:
            return list(DocumentTag.objects.values_list('title', flat=True).order_by('title'))
        except Exception as e:  # noqa: BLE001
            logger.error(f'Error loading tags: {e}')
            return []

    def get_context_data(self, **kwargs) -> dict[str, Any]:
        """Build context data for document list view."""
        context = super().get_context_data(**kwargs)

        # Build queryset through pipeline of operations
        queryset = self._get_base_queryset()
        queryset, current_filter = self._parse_and_apply_filter(queryset)
        queryset, current_sort = self._apply_sorting(queryset)

        # Paginate and add to context
        context['documents'] = self._paginate_queryset(queryset)
        context['current_filter'] = current_filter
        context['current_sort'] = current_sort
        context['available_tags'] = self._get_available_tags()

        return context
